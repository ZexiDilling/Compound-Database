from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook


def export_plate_layout(plate_layout, well_row_col, name, folder):
    wb = Workbook()
    ws = wb.active

    ini_row = 2
    ini_col = 2

    for index_row, row in enumerate(well_row_col["well_row"]):
        for index_col, col in enumerate(well_row_col["well_col"]):
            well = f"{row}{col}"
            colour = plate_layout[well]["colour"].removeprefix("#")

            ws.cell(column=index_col + ini_col, row=index_row + ini_row, value=well).fill \
                = PatternFill(fill_type="solid", fgColor=colour)

    wb.save(f"{folder}/{name}.xlsx")