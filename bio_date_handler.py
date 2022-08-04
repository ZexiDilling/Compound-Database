import configparser
import pandas as pd
from json_handler import dict_reader
import re
from statistics import mean, stdev
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from matplotlib import colors

from bio_data_functions import *


class BIOAnalyser:
    def __init__(self, config, bio_plate_report_setup):
        """
        :param config: the config file for the program
        :type config: configparser.ConfigParser
        :param well_states_report: dict over what state wells should be in, to be printed on the report sheet.
        :type well_states_report: dict
        :param plate_analysis_dict: Method used to analyse the data
        :type plate_analysis_dict: dict
        """

        self.config = config
        self.cal_stuff = {"avg": mean, "stdev": stdev}

        self.well_states_report = bio_plate_report_setup["well_states_report"]
        self.plate_report_calc_dict = bio_plate_report_setup["plate_report_calc_dict"]
        self.plate_calc_dict = bio_plate_report_setup["plate_calc_dict"]
        self.plate_analysis = bio_plate_report_setup["plate_analysis_dict"]
        self.z_prime_calc = bio_plate_report_setup["z_prime_calc"]
        self.heatmap_colours = bio_plate_report_setup["heatmap_colours"]
        self.pora_threshold = bio_plate_report_setup["pora_threshold"]



    def __str__(self):
        """
        A modul that analyse raw data from a plate reader. where the raw data is an excel file.
        The plate reader used for these data is a ????
        :return: the analysed data
        """

    def _plate_well_dict(self):
        """
        Makes a dict over the state of each well (empty, sample, blank...)
        :return: pw_dict
        """
        pw_dict = {}
        for layout in self.plate:
            for counter in self.plate[layout]:
                try:
                    pw_dict[self.plate[layout][counter]["well_id"]] = self.plate[layout][counter]["state"]
                except TypeError:
                    pass

        return pw_dict

    def _data_converter(self, all_data, well_type):
        """
        convert raw data in the analysed data
        :return: all_data, well_row, well_col, pw_dict
        """

        pw_dict = self._plate_well_dict()

        for methode in self.plate_analysis:
            if self.plate_analysis[methode]["used"]:
                self._well_calculations(well_type, all_data, methode)

        all_data["calculations"]["other_data"] = {}
        if self.z_prime_calc:
            all_data["calculations"]["other_data"]["z_prime"] = z_prime(all_data, "normalised")

        return all_data, pw_dict

    def _well_calculations(self, well_type, all_data, methode):
        """
        Calculate each analyse methode for each well
        :param well_type: a dict for each state (empty, sample, blank...) with a list of the wells in that state
        :param all_data: all the data
        :param methode: what analyse method is being used
        :return:
        """
        if methode != "original":
            all_data["plates"][methode] = {}
            all_data["plates"][methode]["wells"] = {}
        for state in well_type:
            all_data["plates"][methode][state] = []
            for well in well_type[state]:
                all_data["plates"][methode]["wells"][well] = self.plate_analysis[methode]["methode"](all_data, well)
                all_data["plates"][methode][state].append(well)

        all_data["calculations"][methode] = {}

        for state in well_type:
            all_data["calculations"][methode][state] = {}
            for calc in self.cal_stuff:
                if self.plate_calc_dict[methode][calc]:
                    try:
                        all_data["calculations"][methode][state][calc] = self.cal_stuff[calc](
                            [all_data["plates"][methode]["wells"][well] for well in all_data["plates"][methode][state]])
                    except ValueError:
                        all_data["calculations"][methode][state][calc] = "NaN"

    def _cal_info(self, ws, init_col, counter_row, temp_dict, methode):
        """
        Writes in the calculation information.
        :param ws: worksheet
        :param init_col: column to writing to
        :param counter_row: a counter for what row to write to
        :param temp_dict: the dict with the data for each well
        :param methode: the analysed method
        :return:counter_row
        """
        temp_row = counter_row
        for state in temp_dict["plates"][methode]:
            temp_col = init_col
            if state != "wells":
                if self.plate_calc_dict[methode]["state"][state]:
                    for calc in temp_dict["calculations"][methode][state]:
                        if counter_row == temp_row:
                            ws[ex_cell(counter_row, temp_col + 1)] = calc
                            ws[ex_cell(counter_row, temp_col + 1)].font = Font(b=True)
                        if temp_col == init_col:
                            ws[ex_cell(counter_row + 1, temp_col)] = state
                            ws[ex_cell(counter_row + 1, temp_col)].font = Font(b=True)
                        ws[ex_cell(counter_row + 1, temp_col + 1)] = temp_dict["calculations"][methode][state][calc]
                        temp_col += 1
                    counter_row += 1
        return counter_row

    def _write_plate(self, ws, counter_row, temp_dict, methode, well_row_col, pw_dict):
        """
        Writes the data for each analyse into the excel file including the calculations
        :param ws: Worksheet
        :param counter_row: what row to write to
        :param temp_dict: the dict for the specific analysed method
        :param methode: what analysed method are being looked at
        :param well_row_col: all the headlines for each row and column
        :param pw_dict: a dict for each well and it's state (empty, sample, blank...)
        :return: counter_row
        """
        indent_col = 3
        indent_row = 3
        init_row = counter_row + indent_row
        init_col = indent_col
        translate_wells_to_cells = {}
        counter_row += indent_row
        for index_row, row in enumerate(well_row_col["well_row"]):

            # sets the headline and colour for the headline for row
            ws.cell(column=-1 + indent_col, row=counter_row, value=row).fill = \
                PatternFill("solid", fgColor="DDDDDD")
            for index_col, col in enumerate(well_row_col["well_col"]):
                if index_row == 0:
                    # Merge cell above tables, and writes the name of the method used for the plate
                    # ws.merged_cells(start_row=counter_row - 2, start_column=indent_col - 1,
                    #                 end_row=counter_row - 2, end_column=indent_col + 1)
                    ws.cell(column=indent_col - 1, row=counter_row - 2, value=methode).font = Font(b=True)

                    # sets the headline and colour for the headline for column
                    ws.cell(column=index_col + indent_col, row=counter_row - 1, value=int(col)).fill = \
                        PatternFill("solid", fgColor="DDDDDD")

                temp_well = f"{row}{col}"
                temp_cell = ex_cell(counter_row, index_col + indent_col)
                translate_wells_to_cells[temp_well] = temp_cell
                # Writes the data in for each well. ignore wells witch state == empty  - - - - -
                #
                #ADD TO SETTINGS!!!
                #
                #
                if temp_well not in temp_dict["plates"][methode]["empty"]:
                    ws.cell(column=index_col + indent_col, row=counter_row,
                            value=temp_dict["plates"][methode]["wells"][temp_well])
            counter_row += 1
        free_col = len(well_row_col["well_col"]) + indent_col

        # Writes the info for the calculation for each method

        if self.plate_calc_dict[methode]["use"]:
            counter_row = self._cal_info(ws, init_col, counter_row, temp_dict, methode)

        # colour wells depending on what state the wells are (sample, blank, min, max...) and add a reading guide.
        if self.plate_analysis[methode]["state_mapping"]:
            state_mapping(self.config, ws, translate_wells_to_cells, self.plate, init_row, free_col, temp_dict, methode)

        # colour in the heat map, if sets to active. Can set for each method
        if self.plate_analysis[methode]["heatmap"]:
            heatmap(self.config, ws, pw_dict, translate_wells_to_cells, self.heatmap_colours)

        if self.plate_analysis[methode]["Hit_Mapping"]:
            hit_mapping(ws, temp_dict, self.pora_threshold, methode, translate_wells_to_cells, free_col, init_row)

        counter_row += 1
        return counter_row

    def cal_writer(self, ws_report, all_data, init_row):
        indent_col = 2
        row_counter = init_row
        for plate_analysed in all_data["calculations"]:
            if self.plate_report_calc_dict[plate_analysed]["use"]:
                ws_report.cell(column=-1 + indent_col, row=row_counter, value=plate_analysed).font = Font(b=True)
                row_counter += 1
                for state in all_data["calculations"][plate_analysed]:
                    try:
                        self.plate_report_calc_dict[plate_analysed]["state"][state]
                        temp_name = "state"
                    except KeyError:
                        temp_name = "calc"

                    if self.plate_report_calc_dict[plate_analysed][temp_name][state]:
                        ws_report.cell(column=indent_col, row=row_counter, value=state).font = Font(b=True)
                        if plate_analysed != "other_data":
                            for calc in all_data["calculations"][plate_analysed][state]:
                                ws_report.cell(column=indent_col + 1, row=row_counter, value=calc)
                                ws_report.cell(column=indent_col + 2, row=row_counter,
                                               value=all_data["calculations"][plate_analysed][state][calc])
                        # for now, only writes z-prime
                        else:
                            if self.plate_report_calc_dict[plate_analysed]["calc"]["z_prime"]:
                                ws_report.cell(column=indent_col + 1, row=row_counter,
                                               value=all_data["calculations"][plate_analysed][state])
                        row_counter += 1
                row_counter += 1
        return ws_report

    def _well_writer(self, ws_report, all_data, init_row):
        """
        Writes Well data from the differen analised method into the report sheet on the excel ark
        :param ws_report: Worksheet - Report
        :param all_data: all the data for each well, and calculations
        :param init_row: row to start writing from in the excel file
        :return: None
        """
        indent_col = 6
        row_counter = init_row
        added = False

        for plate_analysed in all_data["plates"]:
            # Writes headline for data inserts to see where the data is coming from
            ws_report.cell(column=indent_col, row=row_counter, value=plate_analysed).font = Font(b=True)
            row_counter += 1
            for counter in self.plate["well_layout"]:
                for _ in self.plate["well_layout"][counter]:

                    # looks through the plate layout, finds the state for each well and check if it needs to be added
                    # based on bool-statment from well_states_report
                    if self.well_states_report[self.plate["well_layout"][counter]["state"]] and not added:
                        well = self.plate["well_layout"][counter]["well_id"]
                        ws_report.cell(column=indent_col + 1, row=row_counter, value=well)
                        ws_report.cell(column=indent_col + 2, row=row_counter,
                                       value=all_data["plates"][plate_analysed]["wells"][well])
                        added = True
                        row_counter += 1
                added = False
            indent_col += 4
            row_counter = init_row

    def _report_writer_controller(self, wb, all_data):
        """
        pass the data into different modules to write data in to an excel ark
        :param wb: the excel ark / workbook
        :param all_data: all the data for each well, and calculations
        :return: None
        """

        init_row = 2
        ws_report = wb.create_sheet("Report")
        ws_report = self.cal_writer(ws_report, all_data, init_row)
        self._well_writer(ws_report, all_data, init_row)

    def _excel_controller(self, all_data, well_row_col, pw_dict):
        """
        controls the path for the data, to write into an excel file
        :param all_data: all the data for each well, and calculations
        :param well_row_col: all the headlines for each row and column
        :param pw_dict: dict over each well and what state it is (empty, sample, blank....)
        :return: None
        """

        wb = load_workbook(self.ex_file)
        ws_data = wb.create_sheet("analysis")
        counter_row = 0

        # sends each plate-analysed-type into the excel file
        for methode in all_data["plates"]:
            counter_row = self._write_plate(ws_data, counter_row, all_data, methode, well_row_col, pw_dict)
        self._report_writer_controller(wb, all_data)

        wb.save(self.ex_file)

    def bio_data_controller(self, ex_file, plate_layout, all_data, well_row_col, well_type):
        """
        the control modul for the bio analysing
        :return: all_data
        """
        self.ex_file = ex_file
        self.plate = plate_layout

        all_data, pw_dict = self._data_converter(all_data, well_type)
        self._excel_controller(all_data, well_row_col, pw_dict)

        return all_data



if __name__ == "__main__":
    file = "Data_analysis-MTase1-Epigenetic_library_HYCPK5448.xlsx"
    folder = "C:/Users/Charlie/PycharmProjects/structure_search/Bio Data/test_single"
    config = configparser.ConfigParser()
    config.read("config.ini")
    full_file = f"{folder}/{file}"
    plate_file = config["files"]["plate_layouts"]
    try:
        plate_list, archive_plates_dict = dict_reader(plate_file)
    except TypeError:
        plate_list = []
        archive_plates_dict = {}
    plate = archive_plates_dict[plate_list[3]]
    bio_plate_report_setup = {
        "well_states_report": {'sample': True, 'blank': False, 'max': False, 'minimum': False,
                               'positive': False, 'negative': False, 'empty': False},
        "plate_analysis_dict": {"original": {"used": True, "methode": org,
                                             "state_mapping": True, "heatmap": False, "Hit_Mapping": False},
                                "normalised": {"used": True, "methode": norm,
                                               "state_mapping": False, "heatmap": True, "Hit_Mapping": False},
                                "pora": {"used": True, "methode": pora,
                                         "state_mapping": False, "heatmap": False, "Hit_Mapping": True}},
        "z_prime_calc": True,
        "heatmap_colours": {'start': 'light red', 'mid': 'white', 'end': 'light green'},
        "pora_threshold": {"low": {"min": -200,
                                   "max": 0},
                           "mid": {"min": 0,
                                   "max": 30},
                           "high": {"min": 120,
                                    "max": 200},
                           "colour": {"low": "green",
                                      "mid": "yellow",
                                      "high": "blue"}}

    }

    bio_a = BIOAnalyser(config, bio_plate_report_setup)
    bio_a.bio_data_controller(full_file, plate)
